# Written by Yuhang Feng November 2019-April 2020
# Written by Yuhang Feng November 2019-April 2020
# Edited by Roberto Franzosi, Tony May 2022

import sys
import GUI_util
import IO_libraries_util

if IO_libraries_util.install_all_Python_packages(GUI_util.window,"charts_Excel_util",['csv','tkinter','os','collections','openpyxl'])==False:
    sys.exit(0)


import tkinter.messagebox as mb
import openpyxl
from openpyxl import Workbook

import reminders_util

import pandas as pd
import os

import IO_csv_util
import GUI_IO_util
import IO_files_util
import IO_user_interface_util
import charts_util

# ensure filename extension is correct for hover_over effects (xlxm) and no effects (xlsx)
def checkExcel_extension(output_file_name,hover_info_column_list):

    filename, file_extension = os.path.splitext(str(output_file_name))
    if filename!='':
        if len(hover_info_column_list) > 0: # there are hover over effects; must be an xlsm file
            if file_extension != 'xlsm':
                output_file_name = filename + '.xlsm'
        else:
            if file_extension != 'xlsx':
                output_file_name = filename + '.xlsx'
    return str(output_file_name)


# when hover-over data (Labels) are displayed the Excel filename extension MUST be xlsm (for macro VBA enabling)
def prepare_hover_data(inputFilename, hover_info_column, index):
    hover_data = []
    withHeader_var = IO_csv_util.csvFile_has_header(inputFilename) # check if the file has header
    data, headers = IO_csv_util.get_csv_data(inputFilename,withHeader_var) # get the data and header
    if withHeader_var:
        if hover_info_column >= 0:
            hover_data.append([headers[hover_info_column]])
        else:
            hover_data.append(['Hover-over data for series ' + str(index+1)])
    else:
        hover_data.append(['Hover-over data for series ' + str(index+1)])


    for i in range(len(data)):
        if hover_info_column >= 0:
            hover_data.append([data[i][hover_info_column]])
        else:
            hover_data.append([''])
        # print("hover_data",hover_data)
    return hover_data

def get_hover_column_numbers(withHeader_var, headers, hover_info_column_list):
    hover_column_numbers = []

    for i in range(len(
            hover_info_column_list)):  # iterate n times (i.e., len(selected_series), where n is the number of series
        if withHeader_var == 1:
            if hover_info_column_list[i] in headers:
                x = headers.index(hover_info_column_list[i])
            else:
                if len(hover_info_column_list[i]) > 0:
                    mb.showwarning(
                        title='Series No.' + str(i + 1) + ' ' + hover_info_column_list[i] + 'Hover Data Warning',
                        message='The hover-over data column for series No.' + str(
                            i + 1) + ' will be empty.\n\nYou may have entered a column name which does not exist in the input CSV file.\n\nPlease, exit the program, check your input and try again.')
                x = -1
            # y=headers.index(selected_series[i][1])
        else:  # NO headers
            try:
                x = int(hover_info_column_list[i])
            except:
                mb.showwarning(title='Series No.' + str(i + 1) + ' ' + hover_info_column_list[i] + ' Hover Data Header',
                               message='The input csv file has no header so the expected hover-over column header should be numbers(o for A, 1 for B,...) but the ENTERED hover-over data column for series No.' + str(
                                   i + 1) + ' is not a number.\n\nPlease, exit the program, check your input and try again.')
                return
        hover_column_numbers.append(x)
    return hover_column_numbers


#data_to_be_plotted contains the values to be plotted
#   the variable has this format:
#   this includes both headers AND data
#   one series: [[['Name1','Frequency'], ['A', 7]]]
#   two series: [[['Name1','Frequency'], ['A', 7]], [['Name2','Frequency'], ['B', 4]]]
#   three series: [[['Name1','Frequency'], ['A', 7]], [['Name2','Frequency'], ['B', 4]], [['Name3','Frequency'], ['C', 9]]]
#   more series: ..........
#chart_title is the name of the sheet
# the title_series is displayed to the right of the chart as the title of the series
#num_label number of bars, for instance, that will be displayed in a bar chart
#second_y_var is a boolean that tells the function whether a second y axis is needed
#   because it has a different scale and plotted values would otherwise be "masked"
#   ONLY 2 y-axes in a single chart are allowed by openpyxl
#chart_type_list is in form ['line', 'line','bar']... one for each of n series plotted
#when called from scripts other than Excel_charts, the list can be of length 1 although more series may be plotted
#   in which case values are filled below
#output_file_name MUST be of xlsx type, rather tan csv

# when NO hover-over data are displayed the Excel filename extension MUST be xlsx and NOT xlsm (becauuse no macro VBA is enabled in this case)

# returns None if an error is encountered
def create_excel_chart(window,data_to_be_plotted,inputFilename,outputDir,scriptType,
                       chart_title,
                       chart_type_list,
                       column_xAxis_label='',
                       column_yAxis_label='',
                       hover_info_column_list=[],
                       reverse_column_position_for_series_label=False,
                       series_label_list=[],
                       second_y_var=0,
                       second_yAxis_label=''):
    #TODO perhaps all these different imports can be consolidated into a single import?
    if 'pie' in str(chart_type_list).lower():
        from openpyxl.chart import (
                PieChart,
                ProjectedPieChart,
                Reference,
        )
    if 'bar' in str(chart_type_list).lower():
        from openpyxl.chart import (
                BarChart,
                Series,
                Reference,
        )
    if 'line' in str(chart_type_list).lower():
        from openpyxl.chart import (
            LineChart,
            Reference,
            Series,
        )
    if 'scatter' in str(chart_type_list).lower():
        from openpyxl.chart import (
                ScatterChart,
                Reference,
                Series,
        )
    if 'radar' in str(chart_type_list).lower():
        from openpyxl.chart import (
                RadarChart,
                Reference,
        )
    if 'bubble' in str(chart_type_list).lower():
        from openpyxl.chart import (
                BubbleChart,
                Series,
                Reference,
        )
    """
    from the User Manual
    Warning: Openpyxl currently supports chart creation within a worksheet only. Charts in existing workbooks will be lost.
    A single Workbook is saved in a file with extension .xlsx.
    A single Workbook can have multiple Worksheets
    See https://www.pythonexcel.com/openpyxl.php
    """

    # head, tail = os.path.split(chart_outputFilename)
    head, tail = os.path.split(inputFilename)
    num_label=len(data_to_be_plotted[0])-1

    # ValueError: Row numbers must be between 1 and 1048576
    # 1048576 is simply 2 to the 20th power, and thus this number is the largest that can be represented in twenty bits.
    # https://stackoverflow.com/questions/33775423/how-to-set-a-data-type-for-a-column-with-closedxml
    nRecords, nColumns = IO_csv_util.GetNumberOf_Records_Columns_inCSVFile(inputFilename)
    if nRecords > 1048575:
        IO_user_interface_util.timed_alert(window, 2000, 'Warning',
                                           "Excel chart error: The number of rows in the input csv file\n\n" + tail + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.",
                                           False, '', True, '', True)
        # mb.showwarning(title='Excel chart error',message="Excel chart error: The number of rows in the input csv file\n\n" + tail + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.'")
        print("Excel chart error: The number of rows in the input csv file\n\n" + tail + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.")
        return
    if 'bar' in chart_type_list or 'line' in chart_type_list:
        if nRecords > 70:
            IO_user_interface_util.timed_alert(window, 2000, 'Warning',
                                               'The input file\n\n' + inputFilename + '\n\ncontains ' + str(nRecords) + ' records, way too many to be displayed clearly in an Excel line chart.\n\nYOU SHOULD USE PLOTLY WHICH GIVES YOU THE OPTION TO DYNAMICALLY FILTER THE DATA ZOOMING IN ON SPECIFIC DATA SEGMENTS.',
                                               False, '', True, '', True)

    if len(hover_info_column_list) > 0:
        outputExtension='.xlsm'
    else:
        outputExtension = '.xlsx'

    if "NLP" in scriptType and "_" + scriptType + "_" in inputFilename: # do not repeat the same name
        scriptType=''
    chart_outputFilename = IO_files_util.generate_output_file_name(inputFilename, '', outputDir, outputExtension, scriptType, chart_type_list[0],'chart')

    n = len(data_to_be_plotted)

    #while the chart_type_list is complete in the Excel_charts GUI,
    #   when calling this function from other scripts only one chartType is typically passed
    if len(chart_type_list) != n:
        for i in range(n-1):
            chart_type_list.append(chart_type_list[0])

    # TODO unnecessary; creating charts only takes a few seconds
    # startTime = IO_user_interface_util.timed_alert(window, 2000, 'Excel charts', 'Started preparing Excel chart ' + str(chart_type_list) + ' at',
    #                                    True,'Input file: ' + tail,True,'',True)

    # lengths is the list of the number of values for each series (e.g. 5 for series 1, 18 for series 2......)
    # lengths = [5, 18, ......]
    # adding extra lines to series titles to be displayed under the x-axis
    # with 1 series len(lengths) = 1
    lengths = [len(x) for x in
               data_to_be_plotted]  # create a list of length for each list inside in the list data_to_be_plotted(a list contain several lists, each list is a row of output we write in excel)
    if len(lengths) > 3:
        insertLines = '\n\n\n'
    else:
        insertLines = '\n\n'

    # https://stackoverflow.com/questions/51140466/openpyxl-set-color-of-bar-in-bar-chart
    # for color in bar charts depending upon value

    # ensure filename extension is correct for hover_over effects (xlxm) and no effects (xlsx)
    # output_file_name = str(checkExcel_extension(output_file_name,hover_info_column_list))
    if len(hover_info_column_list)>0: # hover-over effects are invoked and the Excel filename extension MUST be xlsm
        if len(chart_type_list)==0:
            mb.showwarning(title='Chart type error', message="No chart type was specified (e.g., line, bubble). The chart could not be created.\n\nPlease, select a chart type and try again!")
            return
        #scriptPath = os.path.dirname(os.path.realpath(__file__))
        fpath = ''
        first_chart_type = chart_type_list[0]
        if chart_type_list and all(type == first_chart_type for type in chart_type_list):
            if first_chart_type=="bar":
                chartName = BarChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "barchartsample.xlsm"
                chartFile = "barchartsample.xlsm"
            elif first_chart_type=="pie":
                chartName = BarChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "piechartsample.xlsm"
                chartFile = "piechartsample.xlsm"
                if len(chart_type_list) > 1:
                    mb.showwarning(title='Pie Chart error', message="If you selected pie chart as the intended chart type for display data, only one group of data can be displayed. The system indicates more than one group of data are selected.\n\nPlease, check your input and try again!")
                    return
            elif first_chart_type=="line":
                chartName = LineChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "linechartsample.xlsm"
                chartFile="linechartsample.xlsm"
            elif first_chart_type=="scatter":
                chartName = ScatterChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "scatterchartsample.xlsm"
                chartFile = "scatterchartsample.xlsm"
                new_data_to_be_plotted = []
                for l in range(len(data_to_be_plotted)):
                    new_data_to_be_plotted.append([])
                    index = 0
                    for i in data_to_be_plotted[l]:
                        index = index + 1
                        try:
                            if index == 1:
                                new_data_to_be_plotted[l].append(i)
                            if index >= 2:
                                x = float(i[0])
                                y = float(i[1])
                                new_data_to_be_plotted[l].append((x,y))
                        except:
                            mb.showwarning(title='Scatter Chart error', message="If you selected a scatter chart as the intended chart type to display data, both X-axis and Y-axis can only contain numeric values. Among the columns selected, at least one contains non-numeric values.\n\nPlease, check your input and try again!")
                            return
                data_to_be_plotted = new_data_to_be_plotted
            else:
                mb.showwarning(title='Chart type error', message="The hover-over feature is only available for Bar, Line, Pie, and Scatter charts. The selected chart type is not allowed.\n\nPlease, check your input and try again!")
                return
        else:
            mb.showwarning(title='Chart type error', message="The hover-over feature for multiple groups of data requires that all  groups have the same chart type. The system indicated more than one chart type.\n\nPlease, check your input and try again!")
            return

        if IO_libraries_util.check_inputPythonJavaProgramFile(chartFile,'lib'+os.sep+'sampleCharts') == 0:
            return

        wb = openpyxl.load_workbook(fpath, read_only=False, keep_vba=True)
        ws1 = wb["Data"]
        ws2 = wb["Labels"]

        # clear data values
        row_count1 = ws1.max_row
        for i in range(row_count1):
            ws1.delete_rows(row_count1 - i)

        # clear Hover-over data (Labels)
        row_count2 = ws2.max_row
        for i in range(row_count2):
            ws2.delete_rows(row_count2 - i)

        if reverse_column_position_for_series_label == True:
            mb.showwarning(title='Reverse Series Label Variable Warning', message="The system indicates that you set reverse var for series labels to be true; however, in the hover-over feature, the series labels can only be the header of the Y-axis values (Column B, C, D,... in 'Data' sheet). Or you can specify series labels in series_label_list.\n\nPlease click 'OK' and continue.")

        for i in range(len(series_label_list)):
            if len(series_label_list[i]) > 0:
                data_to_be_plotted[i][0][1]=series_label_list[i]


        for i in range(max(lengths)): # find the largest length of all series
            row = []
            index = 0
            for stats_list in data_to_be_plotted: # Iterate through all the lists
                if i < len(stats_list): # if i is smaller than the length of the current series
                    tail, tail_noExtension, filename_no_hyperlink = IO_files_util.getFilename(str(stats_list[i][0]))
                    stats_list[i][0] = tail
                    if index > 0:
                        row.append(stats_list[i][1]) # then we append the data
                    else:
                        row += stats_list[i] # then we append the data
                else: # else means the length of current series is smaller than the largest length of all series
                    # lines below are for the situation: in an excel chart, we have multiple series, but they are not the same length.
                    # We append pairs of blank values("") for name, frequency for each row of the series with shorter length
                    # Since we always have two columns for each series (e.g., name, freuency), so there are two append("").
                    row += [""]
                    row += [""]
                index = index + 1
            # fill out data sheet
            ws1.append(row)

        withHeader_var = IO_csv_util.csvFile_has_header(inputFilename) # check if the file has header
        data, headers = IO_csv_util.get_csv_data(inputFilename,withHeader_var) # get the data and header
        hover_column_numbers = get_hover_column_numbers(withHeader_var,headers,hover_info_column_list)
        for i in range(len(hover_column_numbers)):
            hover_data = prepare_hover_data(inputFilename, hover_column_numbers[i], i)
            for j in range(len(hover_data)):
                if j > 1048575:
                    print(
                        "Excel chart error with hover over data: The number of rows in the input csv file\n\n" + inputFilename + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.'\n\nProcessing continues...")
                    break
                else:
                    # fill out Hover-over data (Labels) sheet
                    ws2.cell(row=j + 1, column=i + 1).value = hover_data[j][0]
        names = []
        names.append(chart_title)
        names.append(column_yAxis_label)
        names.append(column_xAxis_label+insertLines)
        for i in range(3):
            # fill out Hover-over data (Labels) sheet
            ws2.cell(row=i+1, column = 26*27).value = names[i]

        reminders_util.checkReminder('*',
                                       reminders_util.title_options_Excel_Charts,
                                       reminders_util.message_Excel_Charts,
                                       True)

    else: # NO hover-over effects; the Excel filename extension MUST be xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws_chart = wb.create_sheet()
        ws_chart.title = "Chart"

        index = 0
        for i in range(max(lengths)): # find the largest length of all series
            row = []
            index = 0
            for stats_list in data_to_be_plotted: # Iterate through all the lists
                # when X-axis values contain a document dressed for hyperlink and with full path
                #   undressed the hyperlink and only display the tail of the document
                if i < len(stats_list): # if i is smaller than the length of the current series
                    tail, tail_noExtension, filename_no_hyperlink = IO_files_util.getFilename(str(stats_list[i][0]))
                    stats_list[i][0] = tail
                    # if index > 0:
                    #     row.append(stats_list[i][1]) # then we append the data
                    # else:
                    #     row += stats_list[i] # then we append the data
                    row += stats_list[i] # then we append the data
                else: # else means the length of current series is smaller than the largest length of all series
                    # lines below are for the situation: in an excel chart, we have multiple series, but they are not the same length.
                    # We append pairs of blank values("") for name, frequency for each row of the series with shorter length
                    # Since we always have two columns for each series (e.g., name, freuency), so there are two append("").
                    row += [""]
                    row += [""]
                index = index + 1
            # fill out data sheet
            ws.append(row)

        #openpyxl only allows a maximum of 2 y axes with different scales
        if second_y_var==0: #we are NOT plotting with 2 y axes with different scales
            chartName=''
            if len(chart_type_list)==0:
                mb.showwarning(title='Chart type error', message="No chart type was specified (e.g., line, bubble). The chart could not be created.\n\nPlease, select a chart type and try again!")
                return
            if str(chart_type_list[0]).lower()=="bar":
                chartName = BarChart()
            elif str(chart_type_list[0]).lower()=="bubble":
                chartName = BubbleChart()
            elif str(chart_type_list[0]).lower()=="line":
                chartName = LineChart()
            elif str(chart_type_list[0]).lower()=="pie":
                chartName = PieChart()
            elif str(chart_type_list[0]).lower()=="radar":
                chartName = RadarChart()
            elif str(chart_type_list[0]).lower()=="scatter":
                chartName = ScatterChart()
            else:
                return
            # Excel allows to group a series value by another series values (e.g., Form or Lemma values by POS or NER tags)
            #   two x-axis labels will be created
            #   https://www.extendoffice.com/documents/excel/2715-excel-chart-group-axis-labels.html
            # but the only way to do this in openpyxl is by plotting TWO separate series,
            #   e.g., a bar chart for Form  or Lemma values and a bar or line chart for POS tags
            #   https://openpyxl.readthedocs.io/en/latest/charts/secondary.html

            if chart_type_list[0].lower()=="line" or chart_type_list[0].lower()=="bar" or chart_type_list[0].lower()=="bubble" or chart_type_list[0].lower()=="scatter":

                if len(column_xAxis_label)>0:
                    chartName.x_axis.title = str(column_xAxis_label)+insertLines
                # else:
                #     chartName.x_axis.title = " X_AXIS"

                if len(column_yAxis_label)>0:
                    chartName.y_axis.title = str(column_yAxis_label) # displayed on the y-axis
                # else:
                #     chartName.y_axis.title = " Y_AXIS"

            if len(series_label_list) > n:
                mb.showwarning(title='Series Label Warning', message="The system indicates that there are more series hover_over_values specified than the number of series (" + str(n) + "). The system will automatically choose the first " + str(n) + " in the series label list.\n\nPlease click 'OK' and continue.")

            for i in range(n): # iterate n times, n is the number of series
                data = Reference(ws, min_col=i*2+2, min_row=2, max_row=1+num_label)
                hover_over_values = Reference(ws,min_col=i*2+1, min_row=2,max_row=1+num_label)

                if chart_type_list[0].lower()=="line" or chart_type_list[0].lower()=="bar" or chart_type_list[0].lower()=="bubble" or chart_type_list[0].lower()=="scatter":
                    if len(series_label_list) > 0 and len(series_label_list[i]) > 0:
                        chartName.series.append(Series(data, title=series_label_list[i]))
                    else:
                        # the title_series is displayed to the right of the chart as the title of the series
                        # should NOT be displayed when you have only one series
                        # for multiple series the series_titles are displayed under the x-axis
                        if reverse_column_position_for_series_label == False:
                            title_series = [t[1] for t in data_to_be_plotted[i]]
                        else:
                            title_series = [t[0] for t in data_to_be_plotted[i]]
                        # title_series is a list [] with two values: title [0] of series and frequency
                        # title_series[0] will be displayed to the right of the chart as the series name
                        #   e.t., Frequencies of NER values
                        # setting title='' will still display a blue little square button w/w series name
                        # chartName.series.append(Series(data)) will still display Series 1
                        # LEGEND
                        # chartName.legend(legendEntry=())
                        # chartName.add_data(data,titles_from_data=False)
                        # test for n > 1 but... no matter how this is exported it still displays the legend oon the right; see comments above
                        if n < 2:
                            chartName.legend=None
                        chartName.series.append(Series(data, title=title_series[0]))
                    chartName.set_categories(hover_over_values)
                else:
                    chartName.add_data(data,titles_from_data=False)
                    chartName.set_categories(hover_over_values)
                chartName.title = chart_title
                if chart_type_list[0].lower()=="line" or chart_type_list[0].lower()=="bar" or chart_type_list[0].lower()=="bubble" or chart_type_list[0].lower()=="scatter":
                    # https://stackoverflow.com/questions/35010050/setting-x-axis-label-to-bottom-in-openpyxl
                    chartName.x_axis.tickLblPos = "low"
                    chartName.x_axis.tickLblSkip = 1 # changing to 2 would skip every other label; 3 every 3; etc.
            ws_chart.add_chart(chartName, "A1")
        else: #plotting with 2 y axes because using different scales
            # if there is no chart at all
            if len(chart_type_list)==0:
                mb.showwarning(title='Chart type error', message="No chart type was specified (e.g., line, bubble). The chart could not be created.\n\nPlease, select a chart type and try again!")
                return
            # if there are more than two charts
            if len(chart_type_list)>2:
                mb.showwarning(title='Number of series error', message="When creating a chart with two y axis, you can ONLY choose two series of data. Here more than two series of data were specified. The chart could not be created.\n\nPlease, select a new pair of series and try again!")
                return

            if chart_type_list[0].lower()=="bar":
                chartName1 = BarChart()
            elif chart_type_list[0].lower()=="bubble":
                chartName1 = BubbleChart()
            elif chart_type_list[0].lower()=="line":
                chartName1 = LineChart()
            elif chart_type_list[0].lower()=="scatter":
                chartName1 = ScatterChart()
            else:
                mb.showwarning(title='Chart type 1 error', message="Wrong chart type selected. Only bar, bubble, line and scatter chart are allowed to have y axis")

            if chart_type_list[1].lower()=="bar":
                chartName2 = BarChart()
            elif chart_type_list[1].lower()=="bubble":
                chartName2 = BubbleChart()
            elif chart_type_list[1].lower()=="line":
                chartName2 = LineChart()
            elif chart_type_list[1].lower()=="scatter":
                chartName2 = ScatterChart()
            else:
                mb.showwarning(title='Chart type 2 error', message="Wrong chart type selected. Only bar, bubble, line and scatter chart are allowed to have y axis")
                return

            # TODO must center the X-axis label
            if len(column_xAxis_label)>0:
                chartName1.x_axis.title = str(column_xAxis_label+insertLines)
            # else:
            #     chartName1.x_axis.title = " X_AXIS"

            if len(column_yAxis_label)>0:
                chartName1.y_axis.title = str(column_yAxis_label)
            # else:
            #     chartName1.y_axis.title = " Y_AXIS"

            #second y-axis label
            if len(second_yAxis_label)>0:
                chartName2.y_axis.title = str(second_yAxis_label)
            # else:
            #     chartName2.y_axis.title = " Second Y_AXIS"


            data = Reference(ws, min_col=2, min_row=2, max_row=1+num_label)
            hover_over_values = Reference(ws,min_col=1, min_row=2,max_row=1+num_label)

            if len(series_label_list) > 2:
                mb.showwarning(title='Series Label Warning', message="The system indicates that there are more series labels specified than the number of series (2). The system will automatically choose the first 2 of the series label list.\n\nPlease click 'OK' and continue.")

            if len(series_label_list) > 0 and len(series_label_list[0]) > 0:
                chartName1.series.append(Series(data, title=series_label_list[0]))
                chartName1.y_axis.title = series_label_list[0]
            else:
                if reverse_column_position_for_series_label == False:
                    title_series = [t[1] for t in data_to_be_plotted[0]]
                else:
                    title_series = [t[0] for t in data_to_be_plotted[0]]

                chartName1.series.append(Series(data, title=title_series[0]))
                chartName1.y_axis.title = title_series[0]
            chartName1.set_categories(hover_over_values)
            chartName1.y_axis.majorGridlines = None

            # Create a second chart
            data = Reference(ws, min_col=4, min_row=2, max_row=1+num_label)
            hover_over_values = Reference(ws,min_col=3, min_row=2,max_row=1+num_label)

            if len(series_label_list) > 0 and len(series_label_list[1]) > 0:
                chartName2.series.append(Series(data, title=series_label_list[1]))
                chartName2.y_axis.title = series_label_list[1]
            else:
                if reverse_column_position_for_series_label == False:
                    title_series = [t[1] for t in data_to_be_plotted[1]]
                else:
                    title_series = [t[0] for t in data_to_be_plotted[1]]
                chartName2.series.append(Series(data, title=title_series[0]))
                chartName2.y_axis.title = title_series[0]
            chartName2.set_categories(hover_over_values)
            chartName2.y_axis.axId = 200
            # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
            chartName1.y_axis.crosses = "max"
            if chart_type_list[0].lower()=="line" or chart_type_list[0].lower()=="bar" or chart_type_list[0].lower()=="bubble" or chart_type_list[0].lower()=="scatter":
                # https://stackoverflow.com/questions/35010050/setting-x-axis-label-to-bottom-in-openpyxl
                chartName1.x_axis.tickLblPos = "low"
                chartName1.x_axis.tickLblSkip = 1  # changing to 2 would skip every other label; 3 every 3; etc.

            chartName1 += chartName2

            ws_chart.add_chart(chartName1, "A1")

    # move chart sheet to first in Excel, so as to open Excel directly on the Chart worksheet rather on the Data worksheet of the Excel chart file
    # https://stackoverflow.com/questions/51082458/move-a-worksheet-in-a-workbook-using-openpyxl-or-xl-or-xlsxwriter
    sheets = wb._sheets
    from_loc = len(sheets) - 1
    to_loc = 0
    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)
    # errorFound=False
    try:
        wb.save(chart_outputFilename)
    except IOError:
        mb.showwarning(title='Output file error', message="Could not write the Excel chart file " + chart_outputFilename + "\n\nA file with the same name is already open. Please close the Excel file and try again!")
        return
    #     errorFound=True
    # if errorFound==True:
    #     chart_outputFilename=''

    # IO_user_interface_util.timed_alert(window, 2000, 'Excel charts', 'Finished preparing Excel chart at',
    #                                    True, '', True, startTime, silent=True)
    return chart_outputFilename

def df_to_list_w_header(df):
    res = []
    header = list(df.columns)
    res.append(header)
    for index, row in df.iterrows():
        temp = [row[tag] for tag in header]
        res.append(temp)
    return res


def df_to_list(df):
    res = []
    header = list(df.columns)
    for index, row in df.iterrows():
        temp = [row[tag] for tag in header]
        res.append(temp)
    return res


def list_to_df(tag_list):
    header = tag_list[0]
    df = pd.DataFrame(tag_list[1:], columns=header)
    return df
