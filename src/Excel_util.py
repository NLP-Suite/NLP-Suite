# Written by Yuhang Feng November 2019-April 2020
# Edited by Roberto Franzosi 

import sys
import GUI_util
import IO_libraries_util

if IO_libraries_util.install_all_packages(GUI_util.window,"Excel_util",['csv','tkinter','os','collections','openpyxl'])==False:
    sys.exit(0)

import tkinter.messagebox as mb
from collections import Counter
import openpyxl
from openpyxl import Workbook

import reminders_util

import pandas as pd
from functools import reduce
import os
import csv

import IO_csv_util
import GUI_IO_util
import IO_files_util
import IO_user_interface_util

# ensure filename extension is correct for hover_over effects (xlxm) and no effects (xlsx)
def checkExcel_extension(output_file_name,hover_info_column_list):

    filename, file_extension = os.path.splitext(str(output_file_name))
    if filename!='':
        if len(hover_info_column_list) > 0: # there are hover over effects; must be an xlsm file
            if file_extension != 'xlsm':
                output_file_name = filename + '.xlsm'
        else:
            if file_extension != 'xlsx':
                output_file_name = filename + '.xlsx'
    return str(output_file_name)

# Prepare the data (data_to_be_plotted) to be used in create_excel_chart with the format:
#   the variable has this format:
# (['little, pig', 22])
#   one serie: [[['Name1','Frequency'], ['A', 7]]]
#   two series: [[['Name1','Frequency'], ['A', 7]], [['Name2','Frequency'], ['B', 4]]]
#   three series: [[['Name1','Frequency'], ['A', 7]], [['Name2','Frequency'], ['B', 4]], [['Name3','Frequency'], ['C', 9]]]
#   more series: ..........
# inputFilename has the full path
# columns_to_be_plotted is a double list [[0, 1], [0, 2], [0, 3]]
# TODO HOW DOES THIS DIFFER FROM def prepare_csv_data_for_chart?
def prepare_data_to_be_plotted(inputFilename, columns_to_be_plotted, chart_type_list,
                               count_var=0, column_yAxis_field_list = []):
    withHeader_var = IO_csv_util.csvFile_has_header(inputFilename) # check if the file has header
    data, headers = IO_csv_util.get_csv_data(inputFilename,withHeader_var) # get the data and header
    if len(data)==0:
        return None
    headers=list(headers)
    count_msg, withHeader_msg = build_timed_alert_message(chart_type_list[0],withHeader_var,count_var)
    if count_var == 1:
        dataRange = get_dataRange(columns_to_be_plotted, data)
        # TODO hover_over labels not being passed, neither are any potential aggregate columns
        #   get_data_to_be_plotted_with_counts is less general than
        data_to_be_plotted = get_data_to_be_plotted_with_counts(inputFilename,withHeader_var,headers,columns_to_be_plotted,column_yAxis_field_list,dataRange)
    else:
        try:
            data = pd.read_csv(inputFilename,encoding='utf-8')
        except:
            try:
                data = pd.read_csv(inputFilename,encoding='ISO-8859-1')
                IO_user_interface_util.timed_alert(window, 2000, 'Warning',
                                                   'Excel-util encountered errors with utf-8 encoding and switched to ISO-8859-1 in reading into pandas the csv file ' + inputFilename)
                print("Excel-util encountered errors with utf-8 encoding and switched to ISO-8859-1 encoding in reading into pandas the csv file " + inputFilename)
            except ValueError as err:
                if 'codec' in str(err):
                    err=str(err) + '\n\nExcel-util encountered errors with both utf-8 and ISO-8859-1 encoding in the function \'prepare_data_to_be_plotted\' while reading into pandas the csv file\n\n' + inputFilename + '\n\nPlease, check carefully the data in the csv file; it may contain filenames with non-utf-8/ISO-8859-1 characters; less likely, the data in the txt files that generated the csv file may also contain non-compliant characters. Run the utf-8 compliance algorithm and, perhaps, run the cleaning algorithm that converts apostrophes.\n\nNO EXCEL CHART PRODUCED.'
                mb.showwarning(title='Input file read error',
                       message=str(err))
                return
        data_to_be_plotted = get_data_to_be_plotted_NO_counts(inputFilename,withHeader_var,headers,columns_to_be_plotted,data)
    return data_to_be_plotted

# called, for instance, from NLP.py
# best approach when all the columns to be plotted are already in the file
#   otherwise, use Excel_util.compute_csv_column_frequencies
# only one hover-over column per series can be selected
# each series plotted has its own hover-over column
#   if the column is the same (e.g., sentence), this must be repeated as many times as there are series 

# columns_to_be_plotted is a double list [[0, 1], [0, 2], [0, 3]] where the first number refers to the x-axis value and the second to the y-axis value
# when count_var=1 the second number gets counted
def run_all(columns_to_be_plotted,inputFilename, outputDir, outputFileLabel,
            chart_type_list,chart_title, column_xAxis_label_var,
            hover_info_column_list=[],
            count_var=0,
            column_yAxis_label_var='Frequencies',
            column_yAxis_field_list = [],
            reverse_column_position_for_series_label=False,
            series_label_list=[], second_y_var=0,second_yAxis_label=''):

    data_to_be_plotted = prepare_data_to_be_plotted(inputFilename,
                                columns_to_be_plotted,
                                chart_type_list,count_var,
                                column_yAxis_field_list)
    if data_to_be_plotted==None:
            return ''
    transform_list = []
    # the following is deciding which type of data is returned from prepare_data_to_be_plotted
    # for the function prepare_data_to_be_plotted branch into two different data handling functions which retruns different data type
    # and due to complexity reasons, we keep them in this way:
    # check the data type for the return value and decide which step to take next
    if not(isinstance(data_to_be_plotted[0], list)):
        for df in data_to_be_plotted:
            header = list(df.columns)
            data = df.values.tolist()
            data.insert(0,header)
            transform_list.append(data)
            data_to_be_plotted=transform_list
    if data_to_be_plotted==None:
            Excel_outputFilename = ""
    else:
        chart_title = chart_title
        Excel_outputFilename = create_excel_chart(GUI_util.window, data_to_be_plotted, inputFilename, outputDir,
                                                  outputFileLabel, chart_title, chart_type_list,
                                                  column_xAxis_label_var, column_yAxis_label_var,
                                                  hover_info_column_list,
                                                  reverse_column_position_for_series_label,
                                                  series_label_list, second_y_var, second_yAxis_label)
    return Excel_outputFilename

#sortOrder = True (descending 3, 2, 1)
#sortOrder = False (ascending 1, 2, 3)
def sort_data (ExcelChartData, sortColumn,sortOrder):
    sorted_data = sorted(ExcelChartData, key=lambda tup:tup[sortColumn],reverse=sortOrder)
    return sorted_data

def build_timed_alert_message(chart_type,withHeader_var,count_var):
    if withHeader_var==1:
        withHeader_msg='WITH HEADERS'
    else:
        withHeader_msg='WITHOUT HEADERS'
    if count_var==1:
        count_msg='WITH COUNTS'
    else:
        count_msg='WITHOUT COUNTS'
    return withHeader_msg, count_msg

# convert the GUI string values for CSV files with or without headers to their column numbers
# selected_series is columns_to_be_plotted (e.g., [1,2]) 
# def get_column_numbers_to_be_plotted(withHeader_var,headers,selected_series):
#     column_numbers_to_be_plotted = []
#     # iterate n times (i.e., len(selected_series),
#     #   where n is the number of series
#
#     for i in range(len(selected_series)):
#         if withHeader_var==1:
#             x=headers.index(selected_series[i][0])
#             y=headers.index(selected_series[i][1])
#         else: #NO headers
#             x = int(selected_series[i][0])-1
#             y = int(selected_series[i][1])-1
#         column_numbers_to_be_plotted.append([x,y])
#     return column_numbers_to_be_plotted

def get_hover_column_numbers(withHeader_var,headers,hover_info_column_list):
    hover_column_numbers = []
    
    for i in range(len(hover_info_column_list)): # iterate n times (i.e., len(selected_series), where n is the number of series
        if withHeader_var==1:
            if hover_info_column_list[i] in headers:
                x=headers.index(hover_info_column_list[i])
            else:
                if len(hover_info_column_list[i]) > 0:
                    mb.showwarning(title='Series No.'+ str(i+1) +' Hover Data Warning', message='The hover-over data column for series No.' + str(i+1) + ' will be empty.\n\nYou may have entered a column name which does not exist in the input CSV file.\n\nPlease, exit the program, check your input and try again.')
                x=-1
            # y=headers.index(selected_series[i][1])
        else: #NO headers
            try:
                x=int(hover_info_column_list[i])
            except:
                mb.showwarning(title='Series No.'+ str(i+1) +' Hover Data Header', message='The input csv file has no header so the expected hover-over column header should be numbers(o for A, 1 for B,...) but the ENTERED hover-over data column for series No.' + str(i+1) + ' is not a number.\n\nPlease, exit the program, check your input and try again.')
                return
        hover_column_numbers.append(x)
    return hover_column_numbers


# split the pairs of gui x y values into two separate lists of x axis values and y axis value
def get_xaxis_yaxis_values(columns_to_be_plotted):
    x = [a[0] for a in columns_to_be_plotted] # select all the x axis number and put them in a list
    y = [a[1] for a in columns_to_be_plotted] # select all the y axis number and put them in a list
    x1 = [ int(b) for b in x ]  # convert them into int type
    y1 = [ int(b) for b in y]  # convert them into int type
    return x1, y1

def get_dataRange(columns_to_be_plotted, data):
    dataRange = []
    for i in range(len(columns_to_be_plotted)):
        for row in data:
            rowValues = list(row[w] for w in columns_to_be_plotted[i])
            dataRange.append(rowValues)
    dataRange = [ dataRange [i:i + len(data)] for i in range(0, len(dataRange), len(data)) ]
    return dataRange

# TODO if hover_over columns are passed, it should concatenate all values, instead of displaying the first one only
#   (e.g. an example run the going UP function in WordNet)
# this function seems to be less general than def compute_csv_column_frequencies; that function handl;es aggregation and hover over effects
# we should consolidate the two and use the most general one under he heading get_data_to_be_plotted_with_counts

# def get_data_to_be_plotted_with_counts(inputFileName,withHeader_var,headers,columns_to_be_plotted,column_yAxis_field_list,dataRange):
#     CALL compute_column_frequencies(columns_to_be_plotted, dataRange, headers,column_yAxis_field_list)
#
#     CALLED compute_column_frequencies(columns_to_be_plotted, data_list, headers,specific_column_value_list=[]):


# -----------------------------------------------------------------
# MUST COMPUTE HOVER OVER VALUES!!! see below

# create a list of unique words to be displayed in hover over
# result = IO_files_util.openCSVFile(outputFilenameCSV1, 'r', 'utf-8')
# DataCaptured = csv.reader(result)
# words = set()
# for row in DataCaptured:
#     words.add(row[0])
# also IO_csv_util.get_csv_field_values(inputfile_name, column_name)

def get_data_to_be_plotted_with_counts(inputFilename,withHeader_var,headers,columns_to_be_plotted,specific_column_value_list,data_list):
    data_to_be_plotted=[]
    # data_to_be_plotted = compute_column_frequencies_4Excel(columns_to_be_plotted, dataRange, headers, column_yAxis_field_list)

    column_list=[]
    column_frequencies=[]
    column_stats=[]
    specific_column_value=''
    complete_column_frequencies=[]
    if len(data_list) != 0:
        for k in range(len(columns_to_be_plotted)):
            res=[]
            if len(specific_column_value_list)>0:
                specific_column_value=specific_column_value_list[k]
            #get all the values in the selected column
            column_list = [i[1] for i in data_list[k]]
            counts = Counter(column_list).most_common()
            if len(headers) > 0:
                id_name_num = columns_to_be_plotted[k][0]
                id_name = headers[id_name_num]
                column_name_num = columns_to_be_plotted[k][1]
                column_name = headers[column_name_num]
                if len(specific_column_value_list)==0:
                    column_frequencies = [[column_name + " values", "Frequencies of " + column_name]]
                else:
                    for y in range(len(specific_column_value_list)):
                        column_frequencies = [[id_name, "Frequencies of " + str(specific_column_value) + " in Column " + str(column_name)]]
            else:
                id_name_num = columns_to_be_plotted[k][0]
                id_name = "column_" + str(id_name_num+1)
                column_name_num = columns_to_be_plotted[k][1]
                column_name = "column_" + str(column_name_num+1)
                if len(specific_column_value)==0:
                    column_frequencies = [[column_name + " values", "Frequencies of " + column_name]]
                else:
                    for y in range(len(specific_column_value_list)):
                        column_frequencies = [[id_name, "Frequencies of " + str(specific_column_value) + " in Column_" + str(column_name_num+1)]]
            if len(specific_column_value) == 0:
                for value, count in counts:
                    column_frequencies.append([value, count])
            else:
                for i in range(len(column_list)):
                    if column_list[i] == specific_column_value:
                        res.append(1)
                    else:
                        res.append(0)
                for j in range(len(data_list[k])):
                    column_frequencies.append([data_list[k][j][0], res[j]])
            data_to_be_plotted.append(column_frequencies)

    return data_to_be_plotted

# [[0,2]], [0], [2]
def get_data_to_be_plotted_NO_counts(inputFilename,withHeader_var,headers,columns_to_be_plotted,data):
    data_to_be_plotted=[]
    for gp in columns_to_be_plotted:
        data.iloc[:, gp[1]].replace('N/A', 0)
        # data.iloc[:, gp[1]].astype('float')
        tempData=data.iloc[:,gp]
        data_to_be_plotted.append(data.iloc[:,gp])
    return data_to_be_plotted

#columns_to_be_plotted is the list of columns pairs (x-axis and y-axis) 
#   where both x-axis is always the same and both correspond to the numeric values of the column position starting at 0
#   it has the format [[0,2],[0,5],...]
#data_list is the datarange we need to compute
#   it has the format 
#   [[['1','5.91'],['2','6.0'],...]]
#headers is all the headers in the original input csv file
#specific_column_value_list contains the list of specific y axis values to be counted, one for each series
#   it has the format ['5.13','13.51','apple'] 
#most_common([n])
#   Return a list of the n most common elements and their counts. 
#   If n is omitted or None, most_common() returns all elements in the counter.
#complete_column_frequencies, the return variable, has the following format of PAIRS of headers & stats is always a frequency...
#   [[['Sentence ID','Frequencies of series 1'],['5',302],...],[['Year','Series 2'],['1981',1000],...]]

# def compute_column_frequencies_4Excel(window, inputFilename, inputDataFrame, output_dir, columns_to_be_plotted,
#                                        select_col, hover_col, group_col, openOutputFiles, createExcelCharts=False,
#                                        fileNameType='CSV', chartType='line', count_var=0):
#     # # outputCsvfilename = IO_util.generate_output_file_name(inputFilename, output_dir, '.csv',fileNameType,'stats')
#     # outputCsvfilename = inputFilename[:-4]+"_stats.csv" #IO_util.generate_output_file_name(inputFilename, output_dir, '.csv',fileNameType,'stats')
#
#     # filesToOpen.append(outputCsvfilename)
#
#     container = []
#     if len(inputDataFrame) != 0:
#         data = inputDataFrame
#     else:
#         with open(inputFilename, encoding='utf-8', errors='ignore') as infile:
#             reader = csv.reader(x.replace('\0', '') for x in infile)
#             headers = next(reader)
#         header_indices = [i for i, item in enumerate(headers) if item]
#         data = pd.read_csv(inputFilename, usecols=header_indices, encoding='utf-8')
#
#     for col in hover_col:
#         temp = group_col.copy()
#         temp.append(col)
#         c = data.groupby(group_col)[col].apply(list).to_dict()
#
#         container.append(c)
#
#     temp = group_col.copy()
#     temp.extend(select_col)
#     data = data.groupby(temp).size().reset_index(name='Frequency')
#     for index, row in data.iterrows():
#         if row[select_col[0]] == '':
#             data.at[index, 'Frequency'] = 0
#
#     hover_header = ', '.join(hover_col)
#     Hover_over_header = ['Hover_over: ' + hover_header]
#     if len(hover_col) != 0:
#         for index, hover in enumerate(hover_col):
#             df = pd.Series(container[index]).reset_index()
#             temp = group_col.copy()
#             temp.append(hover)
#             df.columns = temp
#             data = data.merge(df, how='left', left_on=group_col, right_on=group_col)
#         temp_str = '%s' + '\n%s' * (len(hover_col) - 1)
#         data['Hover_over: ' + hover_header] = data.apply(lambda x: temp_str % tuple(x[h] for h in hover_col), axis=1)
#         data.drop(hover_col, axis=1, inplace=True)
#     return Excel_util.prepare_csv_data_for_chart(window, inputFilename, data, output_dir, select_col,
#                                                                 Hover_over_header, group_col, fileNameType, chartType,
#                                                                 openOutputFiles, createExcelCharts, count_var)


# def multi_group_multi_hover(input, output,group_col ,hover_col):
#     outputFile = IO_util.generate_output_file_name(input,output,'.csv')
#     data = pd.read_csv(input)
#     group_col.extend(hover_col)
#     group = data.groupby(group_col).size().reset_index(name='counts')
#     group.to_csv(outputFile,index = False)
#
# multi_group_multi_hover(os.getcwd()+'/UnitTest/NLP_NVA_NLP_SCNLP_Murphy Miracles thicker than fog CORENLP_CoNLL_NOUN_DEPREL_list.csv',
#                         os.getcwd()+'/UnitTest/output',['DOCUMENT_ID'],['NOUN DEPREL'])

# group_col and hover_col are lists with multiple items, if so wished
# select_col is also a list BUT with one item only
# called from WordNet.py, CoNLL_*, Stanford_CoreNLP_date_annotator
def compute_csv_column_frequencies_NEW(window, inputFileName, inputDataFrame, output_dir, columns_to_be_plotted,
                                       select_col, hover_col, group_col, openOutputFiles, createExcelCharts=False,
                                       fileNameType='CSV', chartType='line', count_var=0):
    # # outputCsvfilename = IO_util.generate_output_file_name(inputFileName, output_dir, '.csv',fileNameType,'stats')
    # outputCsvfilename = inputFileName[:-4]+"_stats.csv" #IO_util.generate_output_file_name(inputFileName, output_dir, '.csv',fileNameType,'stats')

    # filesToOpen.append(outputCsvfilename)

    container = []
    if len(inputDataFrame) != 0:
        data = inputDataFrame
    else:
        with open(inputFileName, encoding='utf-8', errors='ignore') as infile:
            reader = csv.reader(x.replace('\0', '') for x in infile)
            headers = next(reader)
        header_indices = [i for i, item in enumerate(headers) if item]
        data = pd.read_csv(inputFileName, usecols=header_indices, encoding='utf-8')

    for col in hover_col:
        temp = group_col.copy()
        temp.append(col)
        c = data.groupby(group_col)[col].apply(list).to_dict()

        container.append(c)

    temp = group_col.copy()
    temp.extend(select_col)
    data = data.groupby(temp).size().reset_index(name='Frequency')
    for index, row in data.iterrows():
        if row[select_col[0]] == '':
            data.at[index, 'Frequency'] = 0

    hover_header = ', '.join(hover_col)
    Hover_over_header = ['Hover_over: ' + hover_header]
    if len(hover_col) != 0:
        for index, hover in enumerate(hover_col):
            df = pd.Series(container[index]).reset_index()
            temp = group_col.copy()
            temp.append(hover)
            df.columns = temp
            data = data.merge(df, how='left', left_on=group_col, right_on=group_col)
        temp_str = '%s' + '\n%s' * (len(hover_col) - 1)
        data['Hover_over: ' + hover_header] = data.apply(lambda x: temp_str % tuple(x[h] for h in hover_col), axis=1)
        data.drop(hover_col, axis=1, inplace=True)
    return prepare_csv_data_for_chart(window, inputFileName, data, output_dir, select_col,
                                                                Hover_over_header, group_col, fileNameType, chartType,
                                                                openOutputFiles, createExcelCharts, count_var)


# group_col and hover_col are lists with multiple items, if so wished
# select_col is also a list BUT with one item only
# called from WordNet.py, CoNLL_*, Stanford_CoreNLP_date_annotator
def compute_csv_column_frequencies(window, inputFilename, inputDataFrame, outputDir, openOutputFiles, createExcelCharts,
                                        columns_to_be_plotted,
                                        select_col, hover_col, group_col,
                                       fileNameType='CSV', chartType='line', count_var=0):

    # # outputCsvfilename = IO_util.generate_output_file_name(inputFilename, output_dir, '.csv',fileNameType,'stats')
    # outputCsvfilename = inputFilename[:-4]+"_stats.csv" #IO_util.generate_output_file_name(inputFilename, output_dir, '.csv',fileNameType,'stats')

    # filesToOpen.append(outputCsvfilename)

    tempFiles=[]
    container = []

    # TODO remove return when fixed
    return tempFiles

    if len(inputDataFrame) != 0:
        data = inputDataFrame
    else:
        with open(inputFilename, encoding='utf-8', errors='ignore') as infile:
            reader = csv.reader(x.replace('\0', '') for x in infile)
            headers = next(reader)
        header_indices = [i for i, item in enumerate(headers) if item]
        data = pd.read_csv(inputFilename, usecols=header_indices, encoding='utf-8')

    for col in hover_col:
        temp = group_col.copy()
        temp.append(col)
        c = data.groupby(group_col)[col].apply(list).to_dict()

        container.append(c)

    temp = group_col.copy()
    temp.extend(select_col)
    data = data.groupby(temp).size().reset_index(name='Frequency')
    for index, row in data.iterrows():
        if row[select_col[0]] == '':
            data.at[index, 'Frequency'] = 0

    hover_header = ', '.join(hover_col)
    Hover_over_header = ['Hover_over: ' + hover_header]
    if len(hover_col) != 0:
        for index, hover in enumerate(hover_col):
            df = pd.Series(container[index]).reset_index()
            temp = group_col.copy()
            temp.append(hover)
            df.columns = temp
            data = data.merge(df, how='left', left_on=group_col, right_on=group_col)
        temp_str = '%s' + '\n%s' * (len(hover_col) - 1)
        data['Hover_over: ' + hover_header] = data.apply(lambda x: temp_str % tuple(x[h] for h in hover_col), axis=1)
        data.drop(hover_col, axis=1, inplace=True)

        # prepare_data_to_be_plotted(inputFilename, columns_to_be_plotted, chart_type_list, count_var=0,
        #                                column_yAxis_field_list=[]):
        #     withHeader_var = IO_csv_util.csvFile_has_header(inputFilename)  # check if the file has header
        #     data, headers = IO_csv_util.get_csv_data(inputFilename, withHeader_var)  # get the data and header
        #     count_msg, withHeader_msg = build_timed_alert_message(chart_type_list[0], withHeader_var, count_var)
    data.to_csv('freq.csv',index=False)
    if createExcelCharts:
        xAxis='Sentence index'
        chartTitle='Frequency distribution'
        columns_to_be_plotted=[[1,3]]

        Excel_outputFilename=prepare_csv_data_for_chart(window, inputFilename, data, outputDir, select_col, Hover_over_header, group_col,
                                   fileNameType, chartType, openOutputFiles, createExcelCharts, count_var)

        # data_to_be_plotted=prepare_data_to_be_plotted('freq.csv', columns_to_be_plotted,
        #                                      chartType)
        #
        # Excel_outputFilename=create_excel_chart(window, data_to_be_plotted, 'freq.csv', outputDir,
        #                     scriptType=fileNameType,
        #                     chartTitle='',
        #                    chart_type_list=[chartType],
        #                     column_xAxis_label='Sentence index', column_yAxis_label='Frequency',
        #                     hover_info_column_list=[hover_header])

    return Excel_outputFilename #2 files



# generate a new csv file:
# with the set of different values in the select_col ('var') of the input csv file each become a column('var frequency')
# each row of the column representing the frequency of 'var' under the group_by situation of each row.
# the format of the new csv file:
# [groupby_column_1,groupby_column_2, ... ,groupby_column_n, var_1_frequency, var_2_frequency, ... ,var_n_frequency, hover-over]

# hover_col and group_col will be copied from the original file
# called from statistics_csv_util.compute_csv_column_frequencies_NEW

# TODO HOW DOES THIS DIFFER FROM def prepare_data_to_be_plotted?
def prepare_csv_data_for_chart(window,inputfile, inputDataFrame, outputpath, select_col : list, hover_col : list, group_col : list, fileNameType, chartType, openOutputFiles, createExcelCharts,count_var=0):
    filesToOpen=[]
    outputCsvfilename = IO_files_util.generate_output_file_name(inputfile, '', outputpath, '.csv')
    df = inputDataFrame
    # df = pd.read_csv(inputfile)
    # convert a list to a str
    select_column = select_col[0]
    # separate a complete csv file into multiple dataframes filter by select_col, which will produce unequal index numbers
    df_list = sort_by_column(df, select_column)
    # makes those separate dataframes align to the same maximum index
    df_hover = slicing_dataframe(df,group_col + hover_col)
    df_list = align_dataframes(df_list)
    #append aligned dataframes as frequncy columns in the new dataframe
    df_list = [slicing_dataframe(d, group_col + select_col + ['Frequency']) for d in df_list]
    # rename those newly added columns
    df_list = [rename_df(d,select_column) for d in df_list]
    # append the hover-over data in the original csv file
    df_list.append(df_hover)
    # horizontally comcatenate all the frequency dataframes and the hover-over dataframe
    df_merged = reduce(lambda left, right: pd.merge(left, right, how='outer',on=group_col), df_list)
    # replace all the empty strings inside this new df_merged dataframe with 0
    df_merged = df_merged.replace(r'^\s*$', 0, regex=True)
    df_merged.to_csv(outputCsvfilename,index=False) # output
    filesToOpen.append(outputCsvfilename)
    if createExcelCharts:
        columns_to_be_plotted = []
        for i in range(len(df_merged.columns)-1-len(group_col)):
            columns_to_be_plotted.append([1,len(group_col)+i])
        hover_label=[]
        for i in range(len(columns_to_be_plotted)):
            hover_label.append(hover_col[0])
        Excel_outputFilename = run_all(columns_to_be_plotted, inputfile, outputpath,
                                                  'Co-Occ_viewer',
                                                  chart_type_list=[chartType],
                                                  chart_title='Frequency Distribution', column_xAxis_label_var='',
                                                  hover_info_column_list=hover_label,
                                                  count_var=count_var)

        # excel_outputFileName_2 = Excel_util.run_all(columns_to_be_plotted, inputfile, outputpath, outputCsvfilename,
        #                                               chart_type_list=[chartType],
        #                                               chart_title='Frequency Distribution of '+ ','.join(select_col),
        #                                               column_xAxis_label_var='Sentence Index',
        #                                               column_yAxis_label_var='Frequency',
        #                                               outputExtension = '.xlsm', label1=fileNameType,label2=chartType,label3='chart',label4='',label5='', useTime=False,disable_suffix=True,
        #                                               numeric_values_of_column_given = 1,
        #                                             count_var=count_var,
        #                                             column_yAxis_field_list = [],
        #                                             reverse_column_position_for_series_label=False ,
        #                                             series_label_list=[],
        #                                             second_y_var=0,
        #                                             second_yAxis_label='',
        #                                             hover_var=1,
        #                                             hover_info_column_list=hover_label)
        if Excel_outputFilename != "":
            filesToOpen.append(Excel_outputFilename)

    return filesToOpen


# TODO does it compute frequencies by some aggregate values (e.g., document ID)?
# def compute_column_frequencies_4Excel(columns_to_be_plotted, data_list, headers,specific_column_value_list=[]):
#     column_list=[]
#     column_frequencies=[]
#     column_stats=[]
#     specific_column_value=''
#     complete_column_frequencies=[]
#     if len(data_list) != 0:
#         for k in range(len(columns_to_be_plotted)):
#             res=[]
#             if len(specific_column_value_list)>0:
#                 specific_column_value=specific_column_value_list[k]
#             #get all the values in the selected column
#             column_list = [i[1] for i in data_list[k]]
#             counts = Counter(column_list).most_common()
#             if len(headers) > 0:
#                 id_name_num = columns_to_be_plotted[k][0]
#                 id_name = headers[id_name_num]
#                 column_name_num = columns_to_be_plotted[k][1]
#                 column_name = headers[column_name_num]
#                 if len(specific_column_value_list)==0:
#                     column_frequencies = [[column_name + " values", "Frequencies of " + column_name]]
#                 else:
#                     for y in range(len(specific_column_value_list)):
#                         column_frequencies = [[id_name, "Frequencies of " + str(specific_column_value) + " in Column " + str(column_name)]]
#             else:
#                 id_name_num = columns_to_be_plotted[k][0]
#                 id_name = "column_" + str(id_name_num+1)
#                 column_name_num = columns_to_be_plotted[k][1]
#                 column_name = "column_" + str(column_name_num+1)
#                 if len(specific_column_value)==0:
#                     column_frequencies = [[column_name + " values", "Frequencies of " + column_name]]
#                 else:
#                     for y in range(len(specific_column_value_list)):
#                         column_frequencies = [[id_name, "Frequencies of " + str(specific_column_value) + " in Column_" + str(column_name_num+1)]]
#             if len(specific_column_value) == 0:
#                 for value, count in counts:
#                     column_frequencies.append([value, count])
#             else:
#                 for i in range(len(column_list)):
#                     if column_list[i] == specific_column_value:
#                         res.append(1)
#                     else:
#                         res.append(0)
#                 for j in range(len(data_list[k])):
#                     column_frequencies.append([data_list[k][j][0], res[j]])
#             complete_column_frequencies.append(column_frequencies)
#     return complete_column_frequencies


# when hover-over data are displayed the Excel filename extension MUST be xlsm (for macro VBA enabling)
def prepare_hover_data(inputFilename, hover_info_column, index):
    hover_data = []
    withHeader_var = IO_csv_util.csvFile_has_header(inputFilename) # check if the file has header
    data, headers = IO_csv_util.get_csv_data(inputFilename,withHeader_var) # get the data and header
    if withHeader_var:
        if hover_info_column >= 0:
            hover_data.append([headers[hover_info_column]])
        else:
            hover_data.append(['Labels for series ' + str(index+1)])
    else:
        hover_data.append(['Labels for series ' + str(index+1)])


    for i in range(len(data)):
        if hover_info_column >= 0:
            hover_data.append([data[i][hover_info_column]])
        else:
            hover_data.append([''])
        # print("hover_data",hover_data)
    return hover_data
    

#data_to_be_plotted contains the values to be plotted
#   the variable has this format:
#   this includes both headers AND data
#   one serie: [[['Name1','Frequency'], ['A', 7]]]
#   two series: [[['Name1','Frequency'], ['A', 7]], [['Name2','Frequency'], ['B', 4]]]
#   three series: [[['Name1','Frequency'], ['A', 7]], [['Name2','Frequency'], ['B', 4]], [['Name3','Frequency'], ['C', 9]]]
#   more series: ..........
#chartTitle is the name of the sheet
#num_label number of bars, for instance, that will be displayed in a bar chart 
#second_y_var is a boolean that tells the function whether a second y axis is needed 
#   because it has a different scale and plotted values would otherwise be "masked"
#   ONLY 2 y-axes in a single chart are allowed by openpyxl
#chart_type_list is in form ['line', 'line','bar']... one for each of n series plotted
#when called from scripts other than Excel_charts, the list can be of length 1 although more series may be plotted
#   in which case values are filled below
#output_file_name MUST be of xlsx type, rather tan csv

# when NO hover-over data are displayed the Excel filename extension MUST be xlsx and NOT xlsm (becauuse no macro VBA is enabled in this case)

# PREVIOUS
# def create_excel_chart(window,data_to_be_plotted,output_file_name,
#                       chartTitle,
#                       chart_type_list,
#                       column_xAxis_label='',column_yAxis_label='',
#                       reverse_column_position_for_series_label=False,series_label_list=[],second_y_var=0,second_yAxis_label='',inputFilename='',hover_var=0,hover_info_column_list=[]):
def create_excel_chart(window,data_to_be_plotted,inputFilename,outputDir,scriptType,
                       chartTitle,
                       chart_type_list,
                       column_xAxis_label='',
                       column_yAxis_label='',
                       hover_info_column_list=[],
                       reverse_column_position_for_series_label=False,
                       series_label_list=[],
                       second_y_var=0,
                       second_yAxis_label=''):

    #TODO perhaps all these different imports can be consolidated into a single import?
    from openpyxl.chart import (
            PieChart,
            ProjectedPieChart,
            Reference,
    )
    from openpyxl.chart import (
            BarChart,
            Series,
            Reference,
    )
    from openpyxl.chart import (
            LineChart,
            Series,
    )
    from openpyxl.chart import (
            ScatterChart,
            Reference,
            Series,
    )
    from openpyxl.chart import (
            RadarChart,
            Reference,
    )
    from openpyxl.chart import (
            BubbleChart,
            Series,
            Reference,
    )
    """
    from the User Manual
    Warning: Openpyxl currently supports chart creation within a worksheet only. Charts in existing workbooks will be lost.
    A single Workbook is saved in a file with extension .xlsx.
    A single Workbook can have multiple Worksheets
    See https://www.pythonexcel.com/openpyxl.php
    """

    # head, tail = os.path.split(Excel_outputFilename)
    head, tail = os.path.split(inputFilename)
    num_label=len(data_to_be_plotted[0])-1


    # ValueError: Row numbers must be between 1 and 1048576
    # 1048576 is simply 2 to the 20th power, and thus this number is the largest that can be represented in twenty bits.
    # https://stackoverflow.com/questions/33775423/how-to-set-a-data-type-for-a-column-with-closedxml
    if IO_csv_util.GetNumberOfRecordInCSVFile(inputFilename) > 1048575:
        IO_user_interface_util.timed_alert(window, 4000, 'Warning',
                                           "Excel chart error: The number of rows in the input csv file\n\n" + tail + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.'")
        print("Excel chart error: The number of rows in the input csv file\n\n" + tail + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.")
        return

    if len(hover_info_column_list) > 0:
        outputExtension='.xlsm'
    else:
        outputExtension = '.xlsx'

    Excel_outputFilename = IO_files_util.generate_output_file_name(inputFilename, '', outputDir, outputExtension, scriptType, chart_type_list[0],'chart')

    n = len(data_to_be_plotted)

    #while the chart_type_list is complete in the Excel_charts GUI,
    #   when calling this function from other scripts only one chartType is typically passed 
    if len(chart_type_list) != n:
        for i in range(n-1):
            chart_type_list.append(chart_type_list[0])

    IO_user_interface_util.timed_alert(window, 2000, 'Warning', 'Preparing Excel chart ' + tail + '\n\nPlease wait...', False)

    # lenghts is the list of the number of values for each series (e.g. 5 for series 1, 18 for series 2......)
    # lengths = [5, 18, ......]
    lengths = [len(x) for x in
               data_to_be_plotted]  # create a list of length for each list inside in the list data_to_be_plotted(a list contain several lists, each list is a row of output we write in excel)
    if len(lengths) > 3:
        insertLines = '\n\n\n'
    else:
        insertLines = '\n\n'

    # ensure filename extension is correct for hover_over effects (xlxm) and no effects (xlsx)
    # output_file_name = str(checkExcel_extension(output_file_name,hover_info_column_list))
    if len(hover_info_column_list)>0: # hover-over effects are invoked and the Excel filename extension MUST be xlsm
        if len(chart_type_list)==0:
            mb.showwarning(title='Chart type error', message="No chart type was specified (e.g., line, bubble). The chart could not be created.\n\nPlease, select a chart type and try again!")
            return True
        #scriptPath = os.path.dirname(os.path.realpath(__file__))
        fpath = ''
        first_chart_type = chart_type_list[0]
        if chart_type_list and all(type == first_chart_type for type in chart_type_list):
            if first_chart_type=="bar":
                chartName = BarChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "barchartsample.xlsm"
                chartFile = "barchartsample.xlsm"
            elif first_chart_type=="pie":
                chartName = BarChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "piechartsample.xlsm"
                chartFile = "piechartsample.xlsm"
                if len(chart_type_list) > 1:
                    mb.showwarning(title='Pie Chart error', message="If you selected pie chart as the intended chart type for display data, only one group of data can be displayed. The system indicates more than one group of data are selected.\n\nPlease, check your input and try again!")
                    return True
            elif first_chart_type=="line":
                chartName = LineChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "linechartsample.xlsm"
                chartFile="linechartsample.xlsm"
            elif first_chart_type=="scatter":
                chartName = ScatterChart()
                fpath = GUI_IO_util.Excel_charts_libPath + os.sep + "scatterchartsample.xlsm"
                chartFile = "scatterchartsample.xlsm"
                new_data_to_be_plotted = []
                for l in range(len(data_to_be_plotted)):
                    new_data_to_be_plotted.append([])
                    index = 0
                    for i in data_to_be_plotted[l]:
                        index = index + 1
                        try:
                            if index == 1:
                                new_data_to_be_plotted[l].append(i)
                            if index >= 2:
                                x = float(i[0])
                                y = float(i[1])
                                new_data_to_be_plotted[l].append((x,y))
                        except:
                            mb.showwarning(title='Scatter Chart error', message="If you selected a scatter chart as the intended chart type to display data, both X-axis and Y-axis can only contain numeric values. Among the columns selected, at least one contains non-numeric values.\n\nPlease, check your input and try again!")
                            return True
                data_to_be_plotted = new_data_to_be_plotted
            else:
                mb.showwarning(title='Chart type error', message="The hover-over feature is only available for Bar, Line, Pie, and Scatter charts. The selected chart type is not allowed.\n\nPlease, check your input and try again!")
                return True
        else:
            mb.showwarning(title='Chart type error', message="The hover-over feature for multiple groups of data requires that all  groups have the same chart type. The system indicated more than one chart type.\n\nPlease, check your input and try again!")
            return True

        if IO_libraries_util.inputProgramFileCheck(chartFile,'lib'+os.sep+'sampleCharts') == 0:
            return True

        wb = openpyxl.load_workbook(fpath, read_only=False, keep_vba=True)
        ws1 = wb ["Labels"]
        ws2 = wb["Excel data"]

        row_count1 = ws1.max_row
        for i in range(row_count1):
            ws1.delete_rows(row_count1 - i)

        row_count2 = ws2.max_row
        for i in range(row_count2):
            ws2.delete_rows(row_count2 - i)

        if reverse_column_position_for_series_label == True:
            mb.showwarning(title='Reverse Series Label var Warning', message="The system indicates that you set reverse var for series labels to be true; however, in the hover-over feature, the series labels can only be the header of the Y-axis values (Column B, C, D,... in 'Excel data' sheet). Or you can specify series labels in series_label_list.\n\nPlease click 'OK' and continue.")

        for i in range(len(series_label_list)):
            if len(series_label_list[i]) > 0:
                data_to_be_plotted[i][0][1]=series_label_list[i]
           
       
        for i in range(max(lengths)): # find the largest length of all series
            row = []
            index = 0
            for stats_list in data_to_be_plotted: # Iterate through all the lists
                if i < len(stats_list): # if i is smaller than the length of the current series
                    if index > 0:
                        row.append(stats_list[i][1]) # then we append the data
                    else:
                        row += stats_list[i] # then we append the data
                else: # else means the length of current series is smaller than the largest length of all series
                    # Below lines are for the situation: in an excel chart, we have multiple series, but they are not the same length. 
                    # We append pairs of blank values("") for name, frequency for each row of the series with shorter length
                    # Since we always have two columns for each series (e.g., name, freuency), so there are two append("").
                    row += [""]
                    row += [""]
                index = index + 1
            ws2.append(row)

        withHeader_var = IO_csv_util.csvFile_has_header(inputFilename) # check if the file has header
        data, headers = IO_csv_util.get_csv_data(inputFilename,withHeader_var) # get the data and header
        hover_column_numbers = get_hover_column_numbers(withHeader_var,headers,hover_info_column_list)
        for i in range(len(hover_column_numbers)):
            hover_data = prepare_hover_data(inputFilename, hover_column_numbers[i], i)
            for j in range(len(hover_data)):
                if j > 1048575:
                    print(
                        "Excel chart error with hover over data: The number of rows in the input csv file\n\n" + inputFilename + "\n\nexceeds the maximum number of rows Excel can handle (1048576, i.e., 2 to the 20th power, the largest that can be represented in twenty bits), leading to the error 'ValueError: Row numbers must be between 1 and 1048576.'\n\nProcessing continues...")
                    break
                else:
                    ws1.cell(row=j + 1, column=i + 1).value = hover_data[j][0]
        names = []
        names.append(chartTitle)
        names.append(column_yAxis_label)
        names.append(column_xAxis_label+insertLines)
        for i in range(3):
            ws1.cell(row=i+1, column = 26*27).value = names[i]

        reminders_util.checkReminder('*',
                                       reminders_util.title_options_Excel,
                                       reminders_util.message_Excel,
                                       True)

    # NO hover-over effects; the Excel filename extension MUST be xlsx
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws_chart = wb.create_sheet()
        ws_chart.title = "Chart" 

        for i in range(max(lengths)): # find the largest length of all series
            row = []
            for stats_list in data_to_be_plotted: # Iterate through all the lists
                if i < len(stats_list): # if i is smaller than the length of the current series
                    row += stats_list[i] # then we append the data
                else: # else means the length of current series is smaller than the largest length of all series
                    # Below lines are for the situation: in an excel chart, we have multiple series, but they are not the same length. 
                    # We append pairs of blank values("") for name, frequency for each row of the series with shorter length
                    # Since we always have two columns for each series (e.g., name, freuency), so there are two append("").
                    row += [""]
                    row += [""]
            ws.append(row)

        #openpyxl only allows a maximum of 2 y axes with different scales
        if second_y_var==0: #we are NOT plotting with 2 y axes with different scales
            if len(chart_type_list)==0:
                mb.showwarning(title='Chart type error', message="No chart type was specified (e.g., line, bubble). The chart could not be created.\n\nPlease, select a chart type and try again!")
                return True
            if chart_type_list[0]=="bar":
                chartName = BarChart()
            elif chart_type_list[0]=="bubble":
                chartName = BubbleChart()
            elif chart_type_list[0]=="line":
                chartName = LineChart()
            elif chart_type_list[0]=="pie":
                chartName = PieChart()
            elif chart_type_list[0]=="radar":
                chartName = RadarChart()
            elif chart_type_list[0]=="scatter":
                chartName = ScatterChart()
        
            if chart_type_list[0]=="line" or chart_type_list[0]=="bar" or chart_type_list[0]=="bubble" or chart_type_list[0]=="scatter":

                if len(column_xAxis_label)>0:
                    chartName.x_axis.title = column_xAxis_label+insertLines
                # else:
                #     chartName.x_axis.title = " X_AXIS"
                
                if len(column_yAxis_label)>0:
                    chartName.y_axis.title = column_yAxis_label
                # else:
                #     chartName.y_axis.title = " Y_AXIS"

            if len(series_label_list) > n:
                mb.showwarning(title='Series Label Warning', message="The system indicates that there are more series labels specified than the number of series (" + str(n) + "). The system will automatically choose the first " + str(n) + " in the series label list.\n\nPlease click 'OK' and continue.")

            for i in range(n): # iterate n times, n is the number of series
                labels = Reference(ws,min_col=i*2+1, min_row=2,max_row=1+num_label)
                data = Reference(ws, min_col=i*2+2, min_row=2, max_row=1+num_label)

                if chart_type_list[0]=="line" or chart_type_list[0]=="bar" or chart_type_list[0]=="bubble" or chart_type_list[0]=="scatter":
                    if len(series_label_list) > 0 and len(series_label_list[i]) > 0:
                        chartName.series.append(Series(data, title=series_label_list[i]))
                    else:
                        if reverse_column_position_for_series_label == False:
                            title_series = [t[1] for t in data_to_be_plotted[i]]
                        else:
                            title_series = [t[0] for t in data_to_be_plotted[i]]
                        chartName.series.append(Series(data, title=title_series[0]))
                    chartName.set_categories(labels)
                else:
                    chartName.add_data(data,titles_from_data=False)
                    chartName.set_categories(labels)
                chartName.title = chartTitle
                if chart_type_list[0]=="line" or chart_type_list[0]=="bar" or chart_type_list[0]=="bubble" or chart_type_list[0]=="scatter":
                    # https://stackoverflow.com/questions/35010050/setting-x-axis-label-to-bottom-in-openpyxl
                    chartName.x_axis.tickLblPos = "low"
                    chartName.x_axis.tickLblSkip = 2 #
            ws_chart.add_chart(chartName, "A1")
        else: #plotting with 2 y axes because using different scales
            # if there is no chart at all
            if len(chart_type_list)==0:
                mb.showwarning(title='Chart type error', message="No chart type was specified (e.g., line, bubble). The chart could not be created.\n\nPlease, select a chart type and try again!")
                return True
            # if there are more than two charts
            if len(chart_type_list)>2:
                mb.showwarning(title='Number of series error', message="When creating a chart with two y axis, you can ONLY choose two series of data. Here more than two series of data were specified. The chart could not be created.\n\nPlease, select a new pair of series and try again!")
                return True
                
            if chart_type_list[0]=="bar":
                chartName1 = BarChart()
            elif chart_type_list[0]=="bubble":
                chartName1 = BubbleChart()
            elif chart_type_list[0]=="line":
                chartName1 = LineChart()
            elif chart_type_list[0]=="scatter":
                chartName1 = ScatterChart()
            else:
                mb.showwarning(title='Chart type 1 error', message="Wrong chart type selected. Only bar, bubble, line and scatter chart are allowed to have y axis")

            if chart_type_list[1]=="bar":
                chartName2 = BarChart()
            elif chart_type_list[1]=="bubble":
                chartName2 = BubbleChart()
            elif chart_type_list[1]=="line":
                chartName2 = LineChart()
            elif chart_type_list[1]=="scatter":
                chartName2 = ScatterChart()
            else:
                mb.showwarning(title='Chart type 2 error', message="Wrong chart type selected. Only bar, bubble, line and scatter chart are allowed to have y axis")
            
            # TODO must center the X-axis label
            if len(column_xAxis_label)>0:
                chartName1.x_axis.title = str(column_xAxis_label+insertLines)
            # else:
            #     chartName1.x_axis.title = " X_AXIS"
            
            if len(column_yAxis_label)>0:
                chartName1.y_axis.title = str(column_yAxis_label)
            # else:
            #     chartName1.y_axis.title = " Y_AXIS"

            #second y-axis label 
            if len(second_yAxis_label)>0:
                chartName2.y_axis.title = str(second_yAxis_label)
            # else:
            #     chartName2.y_axis.title = " Second Y_AXIS"


            labels = Reference(ws,min_col=1, min_row=2,max_row=1+num_label)
            data = Reference(ws, min_col=2, min_row=2, max_row=1+num_label)

            if len(series_label_list) > 2:
                mb.showwarning(title='Series Label Warning', message="The system indicates that there are more series labels specified than the number of series (2). The system will automatically choose the first 2 of the series label list.\n\nPlease click 'OK' and continue.")

            if len(series_label_list) > 0 and len(series_label_list[0]) > 0:
                chartName.series.append(Series(data, title=series_label_list[0]))
                chartName1.y_axis.title = series_label_list[0]
            else:
                if reverse_column_position_for_series_label == False:
                    title_series = [t[1] for t in data_to_be_plotted[0]]
                else:
                    title_series = [t[0] for t in data_to_be_plotted[0]]

                chartName1.series.append(Series(data, title=title_series[0]))
                chartName1.y_axis.title = title_series[0]
            chartName1.set_categories(labels)
            chartName1.y_axis.majorGridlines = None

            # Create a second chart
            labels = Reference(ws,min_col=3, min_row=2,max_row=1+num_label)
            data = Reference(ws, min_col=4, min_row=2, max_row=1+num_label)

            if len(series_label_list) > 0 and len(series_label_list[1]) > 0:
                chartName.series.append(Series(data, title=series_label_list[1]))
                chartName2.y_axis.title = series_label_list[1]
            else:
                if reverse_column_position_for_series_label == False:
                    title_series = [t[1] for t in data_to_be_plotted[1]]
                else:
                    title_series = [t[0] for t in data_to_be_plotted[1]]
                chartName2.series.append(Series(data, title=title_series[0]))    
                chartName2.y_axis.title = title_series[0]
            chartName2.set_categories(labels)
            chartName2.y_axis.axId = 200
            # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
            chartName1.y_axis.crosses = "max"
            if chart_type_list[0]=="line" or chart_type_list[0]=="bar" or chart_type_list[0]=="bubble" or chart_type_list[0]=="scatter":
                # https://stackoverflow.com/questions/35010050/setting-x-axis-label-to-bottom-in-openpyxl
                chartName.x_axis.tickLblPos = "low"
                chartName.x_axis.tickLblSkip = 2 #

            chartName1 += chartName2

            ws_chart.add_chart(chartName1, "A1")
    errorFound=False
    try:
        wb.save(Excel_outputFilename)
    except IOError:
        mb.showwarning(title='Output file error', message="Could not write the Excel chart file " + Excel_outputFilename + "\n\nA file with the same name is already open. Please close the Excel file and try again!")
        errorFound=True
    if errorFound==True:
        Excel_outputFilename=''
    return Excel_outputFilename

def df_to_list_w_header(df):
    res = []
    header = list(df.columns)
    res.append(header)
    for index, row in df.iterrows():
        temp = [row[tag] for tag in header]
        res.append(temp)
    return res


def df_to_list(df):
    res = []
    header = list(df.columns)
    for index, row in df.iterrows():
        temp = [row[tag] for tag in header]
        res.append(temp)
    return res


def list_to_df(tag_list):
    header = tag_list[0]
    df = pd.DataFrame(tag_list[1:], columns=header)
    return df


def header_check(inputFile):
    sentenceID_pos=''
    docID_pos=''
    docName_pos=''

    if isinstance(inputFile, pd.DataFrame):
        header = list(inputFile.columns)
    else:
        header = IO_csv_util.get_csvfile_headers(inputFile)
    if 'Sentence ID' in header:
        sentenceID_pos = header.index('Sentence ID')
    else:
        pass

    if 'Document ID' in header:
        docID_pos = header.index('Document ID')
    else:
        pass

    if 'Document' in header:
        docName_pos = header.index('Document')
    else:
        pass
    return sentenceID_pos, docID_pos, docName_pos, header


# input can be a csv filename or a dataFrame
# output is a dataFrame
# TODO any funtion that plots data by sentence index shoulld really check that the required sentence IDs are all there and insert them otherwise
def add_missing_IDs(input):
    if isinstance(input, pd.DataFrame):
        df = input
    else:
        df = pd.read_csv(input)
    sentenceID_pos, docID_pos, docName_pos, header = header_check(input)
    Row_list = df_to_list(df)
    for index,row in enumerate(Row_list):
        if index == 0 and Row_list[index][sentenceID_pos] != 1:
            for i in range(Row_list[index][sentenceID_pos]-1,0,-1):
                temp= [''] * len(header)
                for j in range(len(header)):
                    if j == sentenceID_pos:
                        temp[j] = i
                    elif j == docID_pos:
                        temp[j] = Row_list[index][docID_pos]
                    elif j == docName_pos:
                        temp[j] = Row_list[index][docName_pos]
                Row_list.insert(0,temp)
        else:
            if index < len(Row_list)-1 and Row_list[index+1][sentenceID_pos] - Row_list[index][sentenceID_pos] > 1:
                for i in range(Row_list[index+1][sentenceID_pos]-1,Row_list[index][sentenceID_pos],-1):
                    temp = [''] * len(header)
                    for j in range(len(header)):
                        if j == sentenceID_pos:
                            temp[j] = i
                        elif j == docID_pos:
                            temp[j] = Row_list[index][docID_pos]
                        elif j == docName_pos:
                            temp[j] = Row_list[index][docName_pos]
                    Row_list.insert(index+1,temp)
    df = pd.DataFrame(Row_list,columns=header)
    return df


def sort_by_column(input, column):
    if isinstance(input, pd.DataFrame):
        df = input
    else:
        df = pd.read_csv(input)
    col_list = set(df[column].tolist())
    df_list = [df[df[column] == value] for value in col_list]
    return df_list


def align_dataframes(df_list):
    max = 0
    for df in df_list:
        header = list(df.columns)
        if 'Sentence ID' in header:
            sentenceID = 'Sentence ID'
        if df[sentenceID].max() > max:
            max = df[sentenceID].max()
    new_list = []
    for df in df_list:
        if df.empty:
            continue
        temp = df.iloc[-1,:]
        if temp[sentenceID] != max:
            # TODO solve warning issue
            # https://www.dataquest.io/blog/settingwithcopywarning/
            # ​​​​SettingwithCopyWarning
            temp[sentenceID] = max
            temp['Frequency'] = 0
            new_df = df.append(temp,ignore_index=True)
        else:
            new_df = df
        new_list.append(new_df)

    df_list = [add_missing_IDs(data) for data in new_list if not data.empty]
    return df_list


def slicing_dataframe(df,columns):
    df = df[columns]
    return df


def rename_df(df,col):
    for index, row in df.iterrows():
        if row[col] != '':
            name = row[col]
            break
    df.rename(columns={"Frequency": name + " Frequency"},inplace=True)
    df = df.drop(columns=[col])
    return df
